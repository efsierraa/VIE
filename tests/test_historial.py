import base64

from conftest import login

FOTO = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="


def test_paginas_por_seccion_y_rol(client):
    # guarda: dos páginas, cada una con su contenido y su barra de navegación
    login(client, "guarda1")
    ingresos = client.get("/guarda")
    assert ingresos.status_code == 200
    assert "Entrada manual" in ingresos.text
    assert "Registrar paquete" not in ingresos.text  # los paquetes viven en su página
    assert "/guarda/paquetes" in ingresos.text  # enlace en la barra inferior

    paquetes = client.get("/guarda/paquetes")
    assert paquetes.status_code == 200
    assert "Registrar paquete" in paquetes.text
    assert "Pendientes por entregar" in paquetes.text
    assert "Entrada manual" not in paquetes.text
    assert "Iniciar cámara" in paquetes.text  # la entrega lee QR con cámara propia
    assert "reader-pkg" in paquetes.text

    # admin: tres páginas
    login(client, "admin1")
    inicio = client.get("/admin")
    assert inicio.status_code == 200
    assert "paquetes sin residente" in inicio.text
    assert "Nueva cuenta" not in inicio.text

    cuentas = client.get("/admin/cuentas")
    assert cuentas.status_code == 200
    assert "Nueva cuenta" in cuentas.text
    assert "Importar usuarios desde CSV" in cuentas.text

    historial = client.get("/admin/historial")
    assert historial.status_code == 200
    assert "Exportar Excel" in historial.text

    # el residente no entra a las páginas de portería ni administración
    login(client, "residente1")
    assert client.get("/guarda/paquetes").status_code == 403
    assert client.get("/admin/cuentas").status_code == 403
    assert client.get("/admin/historial").status_code == 403


def _datos_de_prueba(client):
    login(client, "residente1")
    r = client.post(
        "/api/visits",
        json={"visitor_name": "Visita Filtrada", "subject": "asunto de visita", "visitor_role": "visitante", "hours": 4},
    )
    assert r.status_code == 200
    login(client, "guarda1")
    r = client.post(
        "/api/packages/manual",
        json={"nombre": "Paquete Filtrado", "description": "caja electronica", "photo_b64": FOTO},
    )
    assert r.status_code == 200


def test_historial_filtros_por_tipo(client):
    _datos_de_prueba(client)
    login(client, "admin1")

    # tipo ingresos: solo visitas
    page = client.get("/admin/historial?tipo=ingresos")
    assert "Visita Filtrada" in page.text
    assert "Paquete Filtrado" not in page.text

    # tipo paquetes: solo paquetes
    page = client.get("/admin/historial?tipo=paquetes")
    assert "Paquete Filtrado" in page.text
    assert "Visita Filtrada" not in page.text

    # ambos: dos tablas
    page = client.get("/admin/historial?tipo=ambos")
    assert "Visita Filtrada" in page.text
    assert "Paquete Filtrado" in page.text

    # el filtro de texto aplica a cada tipo según su campo
    page = client.get("/admin/historial?tipo=ambos&q=caja")
    assert "Paquete Filtrado" in page.text
    assert "Visita Filtrada" not in page.text
    page = client.get("/admin/historial?tipo=ambos&q=Visita")
    assert "Visita Filtrada" in page.text
    assert "Paquete Filtrado" not in page.text

    # cada estado pertenece a su dominio
    page = client.get("/admin/historial?tipo=ambos&estado=pendiente")
    assert "Visita Filtrada" in page.text
    assert "Paquete Filtrado" not in page.text
    page = client.get("/admin/historial?tipo=ambos&estado=en_porteria")
    assert "Paquete Filtrado" in page.text
    assert "Visita Filtrada" not in page.text


def test_export_por_seleccion(client):
    import openpyxl
    from io import BytesIO

    _datos_de_prueba(client)
    login(client, "admin1")

    # sin elegir nada → rechaza
    r = client.get("/admin/exportar")
    assert r.status_code == 400

    # solo ingresos: una sola hoja
    r = client.get("/admin/exportar?ingresos=1")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Ingresos"]

    # solo paquetes: una sola hoja
    r = client.get("/admin/exportar?paquetes=1")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Paquetes"]

    # ambos: las dos hojas
    r = client.get("/admin/exportar?ingresos=1&paquetes=1")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert set(wb.sheetnames) == {"Ingresos", "Paquetes"}

import base64
from datetime import timedelta
from io import BytesIO

from conftest import login

from app.database import SessionLocal
from app.models import Package
from app.routers.api import limpiar_fotos_vencidas
from app.utils import utcnow

FOTO = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="


def _registrar_paquete(client):
    r = client.get("/api/residentes?q=Residenta")
    assert r.status_code == 200
    residentes = r.json()["residentes"]
    assert residentes
    rid = next(x["id"] for x in residentes if x["username"] == "residente1")
    r = client.post(
        "/api/packages",
        json={"resident_id": rid, "description": "caja de prueba", "photo_b64": FOTO},
    )
    assert r.status_code == 200, r.json()
    return r.json()["package"]


def test_flujo_completo_paquete(client):
    login(client, "guarda1")
    pkg = _registrar_paquete(client)
    assert pkg["status"] == "en_porteria"
    code = pkg["short_code"]
    assert len(code) == 6

    # el residente lo ve con foto y QR (el servidor re-codifica todo a JPEG)
    login(client, "residente1")
    r = client.get("/api/packages/mine")
    j = r.json()
    assert j["pendientes"] == 1
    mios = next(p for p in j["packages"] if p["uuid"] == pkg["uuid"])
    assert mios["photo_data_uri"].startswith("data:image/jpeg;base64,")
    assert mios["qr_data_uri"].startswith("data:image/png;base64,")

    r = client.get(f"/api/packages/{pkg['uuid']}/pass")
    assert r.status_code == 200

    # el guarda escanea: ve la foto, aún no marca nada
    login(client, "guarda1")
    r = client.post("/api/packages/scan", json={"code": code.lower()})
    assert r.status_code == 200
    assert r.json()["package"]["photo_data_uri"].startswith("data:image/jpeg;base64,")
    assert r.json()["package"]["status"] == "en_porteria"
    assert r.json()["residente"]["tower"] == "1"

    # entrega: el reloj de borrado de la foto arranca
    r = client.post(f"/api/packages/{pkg['uuid']}/entregar")
    assert r.status_code == 200
    assert r.json()["package"]["status"] == "entregado"
    assert r.json()["package"]["photo_delete_after"] is not None

    # el código ya no sirve
    r = client.post("/api/packages/scan", json={"code": code})
    assert r.status_code == 400

    # el residente confirma la recepción
    login(client, "residente1")
    r = client.post(f"/api/packages/{pkg['uuid']}/confirmar")
    assert r.status_code == 200
    assert r.json()["package"]["status"] == "confirmado"


def test_qr_de_paquete_y_de_visita_no_se_confunden(client):
    from app.security import sign_visit

    login(client, "guarda1")
    pkg = _registrar_paquete(client)
    token_visita = sign_visit(pkg["uuid"])  # sal equivocada a propósito
    r = client.post("/api/packages/scan", json={"token": token_visita})
    assert r.status_code == 400


def test_disputa(client):
    login(client, "guarda1")
    pkg = _registrar_paquete(client)
    client.post(f"/api/packages/{pkg['uuid']}/entregar")
    login(client, "residente1")
    r = client.post(f"/api/packages/{pkg['uuid']}/disputar")
    assert r.status_code == 200
    assert r.json()["package"]["status"] == "disputa"


def test_cancelar_borra_la_foto_de_inmediato(client):
    login(client, "guarda1")
    pkg = _registrar_paquete(client)
    r = client.post(f"/api/packages/{pkg['uuid']}/cancelar")
    assert r.status_code == 200
    assert r.json()["package"]["status"] == "cancelado"
    db = SessionLocal()
    p = db.query(Package).filter(Package.uuid == pkg["uuid"]).first()
    assert p.photo is None
    db.close()


def test_residente_no_registra_ni_escanea(client):
    login(client, "residente1")
    r = client.post("/api/packages", json={"resident_id": 1, "description": "x", "photo_b64": FOTO})
    assert r.status_code == 403
    r = client.post("/api/packages/scan", json={"code": "ABC123"})
    assert r.status_code == 403


def test_otro_residente_no_ve_pase_ajeno(client):
    login(client, "guarda1")
    pkg = _registrar_paquete(client)
    login(client, "residente2")
    r = client.get(f"/api/packages/{pkg['uuid']}/pass")
    assert r.status_code == 404


def test_limpieza_fotos_vencidas(client):
    login(client, "guarda1")
    pkg = _registrar_paquete(client)
    client.post(f"/api/packages/{pkg['uuid']}/entregar")

    # forzar el vencimiento del plazo
    db = SessionLocal()
    p = db.query(Package).filter(Package.uuid == pkg["uuid"]).first()
    assert p.photo is not None
    p.photo_delete_after = utcnow() - timedelta(minutes=1)
    db.commit()
    db.close()

    db = SessionLocal()
    borradas = limpiar_fotos_vencidas(db)
    db.close()
    assert borradas >= 1

    # el registro queda, la foto se libera
    db = SessionLocal()
    p = db.query(Package).filter(Package.uuid == pkg["uuid"]).first()
    assert p.photo is None
    assert p.status == "entregado"
    db.close()


def test_busqueda_de_residentes(client):
    login(client, "guarda1")

    def buscar(q):
        return client.get("/api/residentes", params={"q": q}).json()["residentes"]

    # torre y apto juntos, en todos los formatos
    r = buscar("T1 101")
    assert len(r) == 1 and r[0]["username"] == "residente1"
    assert len(buscar("1 101")) == 1
    assert len(buscar("1-101")) == 1
    assert len(buscar("t1.101")) == 1

    # torre sola o apto solo: no sirven, cero resultados
    assert buscar("1") == []
    assert buscar("101") == []

    # por nombre: apellido solo, o nombres y apellidos juntos
    assert len(buscar("Uno")) == 1
    assert len(buscar("Residenta Uno")) == 1
    assert len(buscar("residente1")) == 1


def test_camara_unificada_reconoce_ambos_qr(client):
    from app.security import sign_package

    # la cámara manda el token a /api/scan/qr: visita → procesa; paquete → muestra para entregar
    login(client, "residente1")
    r = client.post(
        "/api/visits",
        json={"visitor_name": "Visita Cámara", "subject": "x", "visitor_role": "visitante", "hours": 4},
    )
    token_visita = r.json()["token"]

    login(client, "guarda1")
    pkg = _registrar_paquete(client)
    token_paquete = sign_package(pkg["uuid"])

    # QR de visita
    r = client.post("/api/scan/qr", json={"token": token_visita, "action": "entrada"})
    assert r.status_code == 200
    j = r.json()
    assert j["tipo"] == "visita"
    assert j["visit"]["status"] == "dentro"

    # QR de paquete: muestra la foto sin marcar nada
    r = client.post("/api/scan/qr", json={"token": token_paquete, "action": "entrada"})
    assert r.status_code == 200
    j = r.json()
    assert j["tipo"] == "paquete"
    assert j["package"]["photo_data_uri"].startswith("data:image/")
    assert j["package"]["status"] == "en_porteria"

    # entregado → su QR ya no sirve en la cámara
    client.post(f"/api/packages/{pkg['uuid']}/entregar")
    r = client.post("/api/scan/qr", json={"token": token_paquete, "action": "entrada"})
    assert r.status_code == 400


def test_camara_qr_basura(client):
    login(client, "guarda1")
    r = client.post("/api/scan/qr", json={"token": "garbage", "action": "entrada"})
    assert r.status_code == 400
    assert "inválido" in r.json()["detail"]


def test_paquete_tercero_flujo(client):
    login(client, "guarda1")
    r = client.post(
        "/api/packages/manual",
        json={
            "nombre": "Nuevo Vecino",
            "tower": "4",
            "apartment": "1005",
            "description": "paquete de fontanería",
            "photo_b64": FOTO,
        },
    )
    assert r.status_code == 200, r.json()
    pkg = r.json()["package"]
    assert pkg["tercero"] is True
    assert pkg["short_code"]  # QR de reclamo: la cédula sigue siendo la llave de la entrega
    assert pkg["tower"] == "4" and pkg["apartment"] == "1005"  # todo paquete tiene destino
    assert pkg["cedula_tercero"] is None  # la cédula se registra solo al reclamar

    # un residente NO lo ve en su app (los paquetes tercero no tienen dueño con cuenta)
    login(client, "residente1")
    mios = client.get("/api/packages/mine").json()
    assert all(p["uuid"] != pkg["uuid"] for p in mios["packages"])

    # el guarda busca por nombre: ve la foto del paquete
    login(client, "guarda1")
    lista = client.get("/api/packages/terceros", params={"q": "Nuevo"}).json()["paquetes"]
    assert len(lista) == 1
    assert lista[0]["photo_data_uri"].startswith("data:image/")

    # sin cédula no se entrega: se coteja el nombre con la cédula física
    r = client.post(f"/api/packages/{pkg['uuid']}/entregar")
    assert r.status_code == 400
    assert "cédula" in r.json()["detail"]

    # entrega con la cédula de quien reclama, que queda como evidencia
    r = client.post(f"/api/packages/{pkg['uuid']}/entregar", json={"cedula": "1020304050"})
    assert r.status_code == 200
    assert r.json()["package"]["cedula_tercero"] == "1020304050"

    # ya no aparece en la búsqueda (solo en_porteria)
    restantes = client.get("/api/packages/terceros").json()["paquetes"]
    assert all(p["uuid"] != pkg["uuid"] for p in restantes)


def test_paquete_tercero_asignar_a_residente_nuevo(client):
    login(client, "guarda1")
    r = client.post(
        "/api/packages/manual",
        json={"nombre": "Vecino Pendiente", "tower": "5", "apartment": "501", "photo_b64": FOTO},
    )
    pkg = r.json()["package"]

    # administración registra al residente nuevo y asigna el paquete
    login(client, "admin1")
    r = client.post(
        "/api/users",
        json={
            "nombres": "Vecino", "apellidos": "Nuevo", "username": "vecinonuevo",
            "password": "clave123", "role": "residente", "tower": "6", "apartment": "601",
        },
    )
    assert r.status_code == 200
    r = client.post(f"/api/packages/{pkg['uuid']}/asignar", json={"username": "vecinonuevo"})
    assert r.status_code == 200
    j = r.json()["package"]
    assert j["tercero"] is False
    assert j["short_code"]
    assert j["tower"] == "6" and j["apartment"] == "601"  # el destino pasa al dueño real

    # el residente nuevo ya lo ve con QR en su app
    login(client, "vecinonuevo")
    r = client.get("/api/packages/mine")
    assert r.json()["pendientes"] == 1
    assert r.json()["packages"][0]["short_code"] == j["short_code"]
    assert "qr_data_uri" in r.json()["packages"][0]

    # doble asignación no procede
    login(client, "admin1")
    r = client.post(f"/api/packages/{pkg['uuid']}/asignar", json={"username": "vecinonuevo"})
    assert r.status_code == 400


def test_asignar_paquete_tercero_ya_entregado(client):
    """El paquete se entregó a un no registrado; después admin registra al residente
    y vincula el registro con su dueño real (trazabilidad), sin generar QR."""
    login(client, "guarda1")
    r = client.post("/api/packages/manual", json={"nombre": "Entregado a Otro", "tower": "9", "apartment": "909", "photo_b64": FOTO})
    pkg = r.json()["package"]
    client.post(f"/api/packages/{pkg['uuid']}/entregar", json={"cedula": "44332211"})

    login(client, "admin1")
    r = client.post(
        "/api/users",
        json={
            "nombres": "Dueña", "apellidos": "Real", "username": "duenareal",
            "password": "clave123", "role": "residente", "tower": "8", "apartment": "802",
        },
    )
    assert r.status_code == 200
    r = client.post(f"/api/packages/{pkg['uuid']}/asignar", json={"username": "duenareal"})
    assert r.status_code == 200
    j = r.json()["package"]
    assert j["tercero"] is False
    assert j["short_code"]  # conserva su código; ya no tiene QR (entregado)
    assert j["cedula_tercero"] == "44332211"  # la evidencia del reclamo se conserva

    # aparece en el historial ya vinculado: dueña real y evidencia de la cédula del reclamo
    page = client.get("/admin/historial?tipo=paquetes")
    assert "Dueña Real" in page.text
    assert "44332211" in page.text
    assert "Entregado a Otro" not in page.text


def test_paquete_tercero_permisos(client):
    login(client, "residente1")
    r = client.post("/api/packages/manual", json={"nombre": "x", "tower": "1", "apartment": "1", "photo_b64": FOTO})
    assert r.status_code == 403
    r = client.post("/api/packages/abc/asignar", json={"username": "residente1"})
    assert r.status_code == 403
    login(client, "guarda1")
    r = client.post("/api/packages/abc/asignar", json={"username": "residente1"})
    assert r.status_code == 403  # asignar es exclusivo de administración


def test_paquete_tercero_validacion(client):
    login(client, "guarda1")
    # sin torre/apartamento no se registra: todo paquete tiene destino
    r = client.post("/api/packages/manual", json={"nombre": "Alguien", "tower": "", "apartment": "", "photo_b64": FOTO})
    assert r.status_code == 400
    assert "obligatorios" in r.json()["detail"]
    # falta la foto
    r = client.post("/api/packages/manual", json={"nombre": "Alguien", "tower": "4", "apartment": "1005"})
    assert r.status_code == 422


def test_limpieza_foto_tercero_deja_evidencia_de_cedula(client):
    login(client, "guarda1")
    r = client.post(
        "/api/packages/manual",
        json={"nombre": "Evidencia Cedula", "tower": "7", "apartment": "707", "photo_b64": FOTO},
    )
    pkg = r.json()["package"]
    client.post(f"/api/packages/{pkg['uuid']}/entregar", json={"cedula": "556677"})

    db = SessionLocal()
    p = db.query(Package).filter(Package.uuid == pkg["uuid"]).first()
    assert p.photo is not None
    p.photo_delete_after = utcnow() - timedelta(minutes=1)
    db.commit()
    db.close()

    db = SessionLocal()
    limpiar_fotos_vencidas(db)
    p = db.query(Package).filter(Package.uuid == pkg["uuid"]).first()
    assert p.photo is None
    assert p.cedula_tercero == "556677"  # la cédula queda en el registro como evidencia
    db.close()

    db = SessionLocal()
    limpiar_fotos_vencidas(db)
    p = db.query(Package).filter(Package.uuid == pkg["uuid"]).first()
    assert p.photo is None
    assert p.cedula_tercero == "556677"  # la cédula queda en el registro como evidencia
    db.close()


def test_registro_de_entregas_visible_para_cotejo(client):
    # entrega de un paquete tercero: la reclamación se coteja contra este registro
    login(client, "guarda1")
    r = client.post(
        "/api/packages/manual",
        json={"nombre": "Cotejo Vecino", "tower": "9", "apartment": "901", "description": "caja frágil", "photo_b64": FOTO},
    )
    pkg = r.json()["package"]
    client.post(f"/api/packages/{pkg['uuid']}/entregar", json={"cedula": "777888"})

    # el guarda lo ve en el registro de paquetes, con destinatario, cédula y destino
    page = client.get("/guarda/paquetes")
    assert "Cotejo Vecino" in page.text
    assert "777888" in page.text
    assert "caja frágil" in page.text  # la descripción también sirve para cotejar
    assert "T9 · 901" in page.text  # torre y apartamento aunque sea manual
    assert "Ver imagen" in page.text  # la foto sigue vigente (30 días)

    # la foto se sirve por su endpoint, solo para guarda y admin
    foto = client.get(f"/api/packages/{pkg['uuid']}/foto")
    assert foto.status_code == 200
    assert foto.headers["content-type"].startswith("image/jpeg")
    assert foto.content[:2] == b"\xff\xd8"
    login(client, "residente1")
    assert client.get(f"/api/packages/{pkg['uuid']}/foto").status_code == 403

    # administración ve la trazabilidad completa: estado, quién entregó y cuándo
    login(client, "admin1")
    page = client.get("/admin/historial?tipo=paquetes")
    assert "Cotejo Vecino" in page.text
    assert "777888" in page.text
    assert "caja frágil" in page.text
    assert "T9 · 901" in page.text
    assert "entregado" in page.text
    assert "Ver imagen" in page.text
    assert client.get(f"/api/packages/{pkg['uuid']}/foto").status_code == 200


def test_foto_borrada_tras_plazo_muestra_mensaje(client):
    login(client, "guarda1")
    r = client.post(
        "/api/packages/manual",
        json={"nombre": "Foto Vencida", "tower": "3", "apartment": "303", "photo_b64": FOTO},
    )
    pkg = r.json()["package"]
    client.post(f"/api/packages/{pkg['uuid']}/entregar", json={"cedula": "112233"})

    # forzar el vencimiento y pasar la limpieza
    db = SessionLocal()
    p = db.query(Package).filter(Package.uuid == pkg["uuid"]).first()
    p.photo_delete_after = utcnow() - timedelta(minutes=1)
    db.commit()
    db.close()
    db = SessionLocal()
    limpiar_fotos_vencidas(db)
    db.close()

    # el endpoint avisa en lugar de servir la imagen
    r = client.get(f"/api/packages/{pkg['uuid']}/foto")
    assert r.status_code == 404
    assert "no disponible" in r.text
    assert "30 días" in r.text

    # y la tabla muestra el aviso en vez del enlace
    page = client.get("/guarda/paquetes")
    assert "Imagen borrada" in page.text
    assert "Ver imagen" not in page.text.split("Foto Vencida")[1].split("</tr>")[0]


def test_export_incluye_hoja_paquetes(client):
    import openpyxl

    login(client, "guarda1")
    _registrar_paquete(client)
    client.post(
        "/api/packages/manual",
        json={"nombre": "Tercero Excel", "tower": "2", "apartment": "202", "photo_b64": FOTO},
    )
    login(client, "admin1")
    r = client.get("/admin/exportar?ingresos=1&paquetes=1")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert "Paquetes" in wb.sheetnames
    assert "Ingresos" in wb.sheetnames
    hoja = wb["Paquetes"]
    encabezados = [c.value for c in hoja[1]]
    assert "Destinatario" in encabezados and "Cédula" in encabezados